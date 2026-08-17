import io
import logging
from datetime import date
from openpyxl import Workbook
from app.models.animals import Animals, AnimalStatus, Sex
from app.models.vaccinations import Vaccinations
from app.models.inventory import InventoryLot
from app.models.reproduction import ReproductiveEvent, EventType
from app.models.milk_production import MilkProduction
from app.models.financial import Transaction, TransactionType
from app.models.finca import Finca
from app.utils.tenant_context import apply_tenant_filter, get_current_finca_id
from app.utils.financial_filters import exclude_simulated_transactions
from app.extensions import db
from app.services.export_excel_helpers import (
    _make_workbook_response,
    _style_header_row,
    _auto_width,
    _fmt_date,
    _fmt_enum,
)
from app.services.export_pdf_helpers import (
    _build_health_pdf,
    _build_bulk_health_pdf,
    _build_financial_pdf,
)

logger = logging.getLogger(__name__)


class ExportService:
    @staticmethod
    def export_animals_excel(params):
        filters = {}
        if v := params.get("status"):
            try:
                filters["status"] = AnimalStatus(v)
            except ValueError:
                pass
        if v := params.get("sex"):
            try:
                filters["sex"] = Sex(v)
            except ValueError:
                pass
        if v := params.get("breeds_id", type=int):
            filters["breeds_id"] = v

        query = apply_tenant_filter(Animals.query, Animals)
        if filters:
            query = query.filter_by(**filters)
        animals = query.order_by(Animals.record).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Animales"
        ws.append(
            [
                "ID",
                "Código",
                "Sexo",
                "Fecha Nacimiento",
                "Edad (meses)",
                "Peso (kg)",
                "Estado",
                "Raza",
                "ID Padre",
                "ID Madre",
                "Registro",
            ]
        )
        _style_header_row(ws)

        for a in animals:
            ws.append(
                [
                    a.id,
                    a.record,
                    _fmt_enum(a.sex),
                    _fmt_date(a.birth_date),
                    a.age_in_months,
                    a.weight,
                    _fmt_enum(a.status),
                    a.breed.name if a.breed else "",
                    a.father.record if a.father else "",
                    a.mother.record if a.mother else "",
                    _fmt_date(a.created_at),
                ]
            )

        _auto_width(ws)
        return _make_workbook_response(wb, f"animales_{date.today().isoformat()}.xlsx")

    @staticmethod
    def export_vaccinations_excel(params):
        from sqlalchemy import and_

        query = apply_tenant_filter(Vaccinations.query, Vaccinations)
        conditions = []
        if v := params.get("animal_id", type=int):
            conditions.append(Vaccinations.animal_id == v)
        if v := params.get("from_date"):
            try:
                conditions.append(
                    Vaccinations.vaccination_date >= date.fromisoformat(v)
                )
            except ValueError:
                logger.warning("Invalid from_date filter: %s", v)
        if v := params.get("to_date"):
            try:
                conditions.append(
                    Vaccinations.vaccination_date <= date.fromisoformat(v)
                )
            except ValueError:
                logger.warning("Invalid to_date filter: %s", v)
        if conditions:
            query = query.filter(and_(*conditions))
        vacs = query.order_by(Vaccinations.vaccination_date.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Vacunaciones"
        ws.append(
            [
                "ID",
                "Código Animal",
                "Vacuna",
                "Tipo Vacuna",
                "Fecha",
                "Instructor",
                "Aprendiz",
            ]
        )
        _style_header_row(ws)

        for v in vacs:
            ws.append(
                [
                    v.id,
                    v.animals.record if v.animals else "",
                    v.vaccines.name if v.vaccines else "",
                    _fmt_enum(v.vaccines.type) if v.vaccines else "",
                    _fmt_date(v.vaccination_date),
                    v.instructor.fullname if v.instructor else "",
                    v.apprentice.fullname if v.apprentice else "",
                ]
            )

        _auto_width(ws)
        return _make_workbook_response(
            wb, f"vacunaciones_{date.today().isoformat()}.xlsx"
        )

    @staticmethod
    def export_inventory_excel():
        from openpyxl.styles import PatternFill

        lots = (
            apply_tenant_filter(InventoryLot.query, InventoryLot)
            .order_by(InventoryLot.expiry_date)
            .all()
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        ws.append(
            [
                "ID",
                "Tipo",
                "Producto",
                "Lote",
                "Stock Inicial",
                "Stock Actual",
                "Unidad",
                "Vencimiento",
                "Días al vto.",
                "Proveedor",
                "Costo Unit.",
                "Stock Mínimo",
                "Estado",
            ]
        )
        _style_header_row(ws)

        red_fill = PatternFill("solid", fgColor="FFCCCC")
        amber_fill = PatternFill("solid", fgColor="FFF3CC")

        for lot in lots:
            row_idx = ws.max_row + 1
            estado = (
                "Vencido"
                if lot.is_expired
                else ("Stock bajo" if lot.is_low_stock else "OK")
            )
            ws.append(
                [
                    lot.id,
                    _fmt_enum(lot.product_type),
                    lot.product_name or "",
                    lot.lot_number,
                    lot.quantity,
                    lot.current_quantity,
                    lot.unit,
                    _fmt_date(lot.expiry_date),
                    lot.days_to_expiry,
                    lot.supplier or "",
                    lot.unit_cost,
                    lot.min_stock,
                    estado,
                ]
            )
            fill = (
                red_fill
                if lot.is_expired
                else (
                    amber_fill
                    if lot.is_low_stock
                    or (lot.days_to_expiry is not None and lot.days_to_expiry <= 30)
                    else None
                )
            )
            if fill:
                for cell in ws[row_idx]:
                    cell.fill = fill

        _auto_width(ws)
        return _make_workbook_response(
            wb, f"inventario_{date.today().isoformat()}.xlsx"
        )

    @staticmethod
    def export_reproduction_excel(params):
        query = apply_tenant_filter(ReproductiveEvent.query, ReproductiveEvent)
        if v := params.get("animal_id", type=int):
            query = query.filter_by(animal_id=v)
        if v := params.get("event_type"):
            try:
                query = query.filter_by(event_type=EventType(v))
            except ValueError:
                logger.warning("Invalid event_type filter: %s", v)
        events = query.order_by(ReproductiveEvent.event_date.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reproducción"
        ws.append(
            [
                "ID",
                "Hembra",
                "Tipo Evento",
                "Fecha",
                "Macho",
                "Técnica",
                "Resultado Diagnóstico",
                "Fecha Parto Esperada",
                "Días al Parto",
                "Crías Vivas",
                "Crías Muertas",
                "Complicaciones",
                "Notas",
            ]
        )
        _style_header_row(ws)

        for ev in events:
            ws.append(
                [
                    ev.id,
                    ev.animal.record if ev.animal else "",
                    _fmt_enum(ev.event_type),
                    _fmt_date(ev.event_date),
                    ev.sire.record if ev.sire else "",
                    _fmt_enum(ev.technique),
                    _fmt_enum(ev.diagnosis_result),
                    _fmt_date(ev.expected_birth_date),
                    ev.days_to_birth,
                    ev.alive_count,
                    ev.dead_count,
                    "Sí"
                    if ev.complications
                    else ("No" if ev.complications is not None else ""),
                    ev.notes or "",
                ]
            )

        _auto_width(ws)
        return _make_workbook_response(
            wb, f"reproduccion_{date.today().isoformat()}.xlsx"
        )

    @staticmethod
    def export_animal_health_pdf(animal_id):
        from app.models.treatments import Treatments
        from app.models.animalDiseases import AnimalDiseases
        from app.models.control import Control

        animal = Animals.get_by_id(animal_id)
        if not animal:
            return None, "Animal no encontrado"

        controls = (
            Control.query.filter_by(animal_id=animal_id)
            .order_by(Control.checkup_date.desc())
            .all()
        )
        vaccinations = (
            Vaccinations.query.filter_by(animal_id=animal_id)
            .order_by(Vaccinations.vaccination_date.desc())
            .all()
        )
        treatments = (
            Treatments.query.filter_by(animal_id=animal_id)
            .order_by(Treatments.treatment_date.desc())
            .all()
        )
        diseases = (
            AnimalDiseases.query.filter_by(animal_id=animal_id)
            .order_by(AnimalDiseases.diagnosis_date.desc())
            .all()
        )

        pdf_bytes = _build_health_pdf(
            animal, controls, vaccinations, treatments, diseases
        )
        fname = f"historial_{animal.record}_{date.today().isoformat()}.pdf"
        return io.BytesIO(pdf_bytes), fname

    @staticmethod
    def export_bulk_health_pdf(animal_ids_str):
        if animal_ids_str == "all":
            finca_id = get_current_finca_id()
            animal_ids = [
                a.id for a in Animals.query.filter_by(finca_id=finca_id).all()
            ]
        else:
            try:
                animal_ids = [
                    int(i.strip()) for i in animal_ids_str.split(",") if i.strip()
                ]
            except (ValueError, TypeError):
                return (
                    None,
                    "animal_ids debe ser una lista de enteros separados por coma o 'all'",
                )

        if not animal_ids:
            return None, "No se encontraron animales para exportar"

        pdf_bytes = _build_bulk_health_pdf(animal_ids)
        return io.BytesIO(pdf_bytes), f"reporte_lote_{date.today().isoformat()}.pdf"

    @staticmethod
    def export_milk_production_excel(params):
        query = apply_tenant_filter(MilkProduction.query, MilkProduction)
        if v := params.get("animal_id", type=int):
            query = query.filter_by(animal_id=v)
        if v := params.get("from_date"):
            try:
                query = query.filter(MilkProduction.date >= date.fromisoformat(v))
            except ValueError:
                logger.warning("Invalid milk from_date filter: %s", v)
        if v := params.get("to_date"):
            try:
                query = query.filter(MilkProduction.date <= date.fromisoformat(v))
            except ValueError:
                logger.warning("Invalid milk to_date filter: %s", v)

        records = query.order_by(MilkProduction.date.desc()).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Produccion_Leche"
        ws.append(
            [
                "ID",
                "Fecha",
                "Animal",
                "Litros",
                "Sesión",
                "Grasa %",
                "Proteína %",
                "Cél. Somáticas",
                "Notas",
            ]
        )
        _style_header_row(ws)

        for r in records:
            ws.append(
                [
                    r.id,
                    _fmt_date(r.date),
                    r.animals.record
                    if hasattr(r, "animals") and r.animals
                    else r.animal_id,
                    r.liters,
                    _fmt_enum(r.milking_session),
                    r.fat_percentage,
                    r.protein_percentage,
                    r.somatic_cells,
                    r.notes or "",
                ]
            )

        _auto_width(ws)
        return _make_workbook_response(
            wb, f"produccion_leche_{date.today().isoformat()}.xlsx"
        )

    @staticmethod
    def export_financials_excel(params):
        query = exclude_simulated_transactions(
            apply_tenant_filter(Transaction.query, Transaction).filter(
                Transaction.is_deleted.is_(False)
            ),
            Transaction,
        )
        if v := params.get("type"):
            query = query.filter(Transaction.transaction_type == v)
        if v := params.get("category"):
            query = query.filter(Transaction.category == v)

        txs = query.order_by(Transaction.date.desc()).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Finanzas"
        ws.append(
            ["ID", "Fecha", "Tipo", "Categoría", "Monto", "Animal", "Descripción"]
        )
        _style_header_row(ws)

        for t in txs:
            ws.append(
                [
                    t.id,
                    _fmt_date(t.date),
                    _fmt_enum(t.transaction_type),
                    _fmt_enum(t.category),
                    t.amount,
                    t.animal.record if t.animal else "General",
                    t.description or "",
                ]
            )

        _auto_width(ws)
        return _make_workbook_response(wb, f"finanzas_{date.today().isoformat()}.xlsx")

    @staticmethod
    def export_financial_pdf():
        finca_id = get_current_finca_id()
        finca = Finca.query.get(finca_id)
        txs = (
            exclude_simulated_transactions(
                apply_tenant_filter(Transaction.query, Transaction).filter(
                    Transaction.is_deleted.is_(False)
                ),
                Transaction,
            )
            .order_by(Transaction.date.desc())
            .all()
        )
        income = sum(
            [t.amount for t in txs if t.transaction_type == TransactionType.Income]
        )
        expenses = sum(
            [t.amount for t in txs if t.transaction_type == TransactionType.Expense]
        )

        pdf_bytes = _build_financial_pdf(finca, txs, income, expenses)
        return io.BytesIO(pdf_bytes), f"reporte_financiero_{finca.name}.pdf"

    @staticmethod
    def export_multi_finca_general_pdf(user_id):
        from app.models.user import User

        user = User.query.get(user_id)
        user_name = user.fullname if user else "Usuario"

        # Mismo cálculo que la vista panorámica: un solo origen evita que el
        # PDF y la pantalla muestren totales distintos.
        from app.services.finca_kpis import get_user_fincas_report

        fincas_data = [
            {
                "name": row["finca_name"],
                "type": row["finca_type"] or "Sin tipo",
                "location": f"{row['municipality']}, {row['department']}".strip(", "),
                "role": row["role"] or "Sin rol",
                "total_animals": row["kpis"]["total_animals"],
                "total_milk": row["kpis"]["total_milk_liters"],
                "total_income": row["kpis"]["total_income"],
                "total_expenses": row["kpis"]["total_expenses"],
                "net_balance": row["kpis"]["net_balance"],
            }
            for row in get_user_fincas_report(user_id)
        ]

        from app.services.export_pdf_helpers import _build_multi_finca_general_pdf

        pdf_bytes = _build_multi_finca_general_pdf(fincas_data, user_name)
        return io.BytesIO(
            pdf_bytes
        ), f"reporte_general_multi_finca_{date.today().isoformat()}.pdf"

    @staticmethod
    def export_finca_detail_pdf(user_id, finca_id):
        from app.models.user import User
        from app.models.user_finca import UserFinca
        from app.models.finca import Finca
        from app.models.milk_production import MilkProduction
        from app.models.financial import Transaction
        from sqlalchemy import func

        # Validar acceso
        if not UserFinca.has_access(user_id, finca_id):
            return None, "No tiene acceso a esta finca"

        finca = Finca.query.get(finca_id)
        if not finca or finca.is_deleted:
            return None, "Finca no encontrada"

        user = User.query.get(user_id)
        user_name = user.fullname if user else "Usuario"

        # Mismos KPIs que la vista panorámica (excluyen filas con soft delete).
        from app.services.finca_kpis import get_fincas_kpis

        kpis = get_fincas_kpis([finca_id])[finca_id]

        # Promedio por registro de ordeño, no promedio diario.
        avg_milk = (
            db.session.query(func.avg(MilkProduction.liters))
            .filter(
                MilkProduction.finca_id == finca_id,
                MilkProduction.is_deleted.is_(False),
            )
            .scalar()
            or 0.0
        )

        recent_txs = (
            Transaction.query.filter(
                Transaction.finca_id == finca_id, Transaction.is_deleted.is_(False)
            )
            .order_by(Transaction.date.desc())
            .limit(5)
            .all()
        )
        txs_data = [
            {
                "date": t.date,
                "type": t.transaction_type.value,
                "category": t.category.value,
                "amount": float(t.amount),
                "description": t.description or "",
            }
            for t in recent_txs
        ]

        stats = {
            "total_animals": kpis["total_animals"],
            "total_males": kpis["total_animals_males"],
            "total_females": kpis["total_animals_females"],
            "total_milk": kpis["total_milk_liters"],
            "avg_milk": float(avg_milk),
            "total_income": kpis["total_income"],
            "total_expenses": kpis["total_expenses"],
            "net_balance": kpis["net_balance"],
            "total_fields": kpis["total_fields"],
            "total_area": kpis["total_fields_area"],
            "recent_transactions": txs_data,
        }

        from app.services.export_pdf_helpers import _build_finca_detail_pdf

        pdf_bytes = _build_finca_detail_pdf(finca, stats, user_name)
        return (
            io.BytesIO(pdf_bytes),
            f"reporte_finca_{finca.name.replace(' ', '_')}_{date.today().isoformat()}.pdf",
        )
