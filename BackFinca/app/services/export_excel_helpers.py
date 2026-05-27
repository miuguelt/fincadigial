"""Excel export helpers for exports_namespace."""
import io
from datetime import date, datetime

import flask


def _make_workbook_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return flask.send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _style_header_row(ws, header_fill_hex='1F6B35'):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill('solid', fgColor=header_fill_hex)
    font = Font(bold=True, color='FFFFFF')
    border_side = Side(style='thin', color='DDDDDD')
    border = Border(bottom=border_side)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, min_w), max_w)


def _fmt_date(val):
    if val is None:
        return ''
    if isinstance(val, (date, datetime)):
        return val.strftime('%d/%m/%Y')
    return str(val)


def _fmt_enum(val):
    if val is None:
        return ''
    return val.value if hasattr(val, 'value') else str(val)

